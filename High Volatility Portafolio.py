import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

stocks = ['SPCE', 'SAVA', 'RIOT', 'TSLA', 'AMD', 'ON'
          , 'XPEV', 'WISH', 'NIO', 'AFRM', 'AMAT',
           'NVDA', 'SQ',  'MU',  'BABA' ]

data = yf.download(stocks, start='2023-10-19', end='2024-04-19')

adj_close = data['Adj Close']
returns = np.log(adj_close / adj_close.shift(1)).dropna()

mean_returns = returns.mean()
std_dev = returns.std()
annual_mean_returns = returns.mean() * 250
annual_std_dev = returns.std() * np.sqrt(250)

wb = Workbook()

ws1 = wb.create_sheet('AdjClose')
for r in dataframe_to_rows(adj_close, index=True, header=True):
    ws1.append(r)


ws2 = wb.create_sheet('Returns')
for r in dataframe_to_rows(returns, index=True, header=True):
    ws2.append(r)

ws3 = wb.create_sheet('Maximization')


ws3['A1'] = 'Stocks'
ws3['A2'] = 'daily mean_returns'
ws3['A3'] = 'daily std_dev'
ws3['A4'] = 'annual mean_returns'
ws3['A5'] = 'annual std_dev'

for i, stock in enumerate(stocks, start=2):
    ws3.cell(row=1, column=i, value=stock)
    ws3.cell(row=2, column=i, value=mean_returns[stock])
    ws3.cell(row=3, column=i, value=std_dev[stock])
    ws3.cell(row=4, column=i, value=annual_mean_returns[stock])
    ws3.cell(row=5, column=i, value=annual_std_dev[stock])

for i in range(1, len(stocks) + 1):
    ws3.cell(row=7, column=i+1, value=f"w{i}")



ws3.cell(row=7, column=len(stocks)+2, value="x2")

ws3['B10'] = 'n'
ws3['B11'] = 'δ'
ws3['B12'] = 'ε1'
ws3['B13'] = 'ε2'
ws3['B14'] = 'r'

ws3['C10'] = len(stocks)
ws3['C11'] = 1000000
ws3['C12'] = 0.8
ws3['C13'] = 0.2
ws3['C14'] = 0.05


ws3['F10'] = 'max'

ws3['F12'] = 's.a.'
ws3.cell(row=12, column=7, value=f"=SUM(B8:{chr(66+len(stocks))}8)") 
ws3.cell(row=13, column=7, value=f"=SUM(B8:{chr(65+len(stocks))}8)") 
ws3.cell(row=14, column=7, value=f"={chr(66+len(stocks))}8") 
ws3.cell(row=15, column=7, value=f"= {len(stocks)-1} * (SUM(B8:{chr(64+len(stocks))}8) - {chr(65+len(stocks))}8)") 

ws3['H12'] = '≤'
ws3['H13'] = '≤'
ws3['H14'] = '≥'
ws3['H15'] = '≥'

ws3['I12'] = '=C11'
ws3['I13'] = '=C11 * C12'
ws3['I14'] = '=C11 * C13'
ws3['I15'] = 0

ws3.cell(row=10, column=7, value=f"=SUMPRODUCT(B8:{chr(65+len(stocks))}8, B4:{chr(65+len(stocks))}4)+ {chr(66+len(stocks))}8* C14")

wb.remove(wb['Sheet'])
wb.save('Stocks.xlsx')